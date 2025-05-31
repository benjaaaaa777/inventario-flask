from flask import Flask, render_template, request, redirect, send_file
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__)
excel_file = "Formato_Inventario_Proyecto.xlsx"

@app.route('/', methods=['GET', 'POST'])
def index():
    df = pd.read_excel(excel_file, engine='openpyxl', header=1)

    if request.method == 'POST':
        nueva_fila = [request.form.get(col) for col in df.columns]

        wb = load_workbook(excel_file)
        ws = wb.active

        fila_inicio = 3  # empieza desde la fila 3
        # Encuentra la primera fila vacía desde la fila 3
        while True:
            celdas = [ws.cell(row=fila_inicio, column=col).value for col in range(1, ws.max_column + 1)]
            if all(c in (None, "") for c in celdas):
                break
            fila_inicio += 1

        # Escribir nueva fila exactamente allí
        for col_idx, valor in enumerate(nueva_fila, 1):
            ws.cell(row=fila_inicio, column=col_idx).value = valor


        wb.save(excel_file)
        return redirect('/')

    df = df.dropna(how='all')
    return render_template("index.html", data=df.to_dict(orient='records'), columns=df.columns)

@app.route('/descargar')
def descargar():
    return send_file(excel_file, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
